# -*- coding: utf-8 -*-
"""
"Islom aka.xlsx" kontaktlarini MUKAMMAL tozalaydi.

Bajaradigan ishlari:
  1) Kirill harflarni lotinga o'giradi   (Нурметов -> Nurmetov)
  2) Apostroflarni birxillashtiradi      (oʻgʻli / o‘g‘li / o'gli -> o'g'li)
  3) Ismlarni to'g'ri bosh harf bilan formatlaydi
  4) Telefonni bitta formatga keltiradi  (+998XXXXXXXXX)
       - 9 / 10 (8-trunk) / 11 / 12 / 18 (ikkita raqam) xonali holatlar
       - operator prefiksini tekshiradi
       - soxta raqamlarni (hammasi bir xil raqam) aniqlaydi
  5) Bitta katakdagi 2-raqamni ham saqlaydi (yo'qotmaydi)
  6) Dublikat raqamlarni olib tashlaydi
  7) Tuzatib bo'lmaydiganlarni "Tekshirish kerak" varag'iga ajratadi

Natija:  Islom_aka_toza.xlsx  (Toza | Tekshirish kerak | Hisobot)

Ishga tushirish:
    .venv\\Scripts\\python.exe tozalash.py
"""
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

IN_FILE = "Islom aka.xlsx"
OUT_FILE = "Islom_aka_toza.xlsx"

# O'zbekiston mobil operator prefikslari
VALID_PREFIX = {"90", "91", "93", "94", "95", "97", "98", "99", "88", "33", "77", "50", "20", "55"}

# Kirill -> Lotin (o'zbek + rus)
CYR = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo", "ж": "j", "з": "z",
    "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
    "с": "s", "т": "t", "у": "u", "ф": "f", "х": "x", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sh",
    "ъ": "'", "ы": "i", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    "ў": "o'", "қ": "q", "ғ": "g'", "ҳ": "h", "ё": "yo",
}
APOSTROPHES = "ʻʼ‘’`´′"   # turli apostrof belgilar -> '


def translit(s):
    out = []
    for ch in s:
        low = ch.lower()
        if low in CYR:
            rep = CYR[low]
            out.append(rep.upper() if ch.isupper() and rep else rep)
        else:
            out.append(ch)
    return "".join(out)


def clean_name(raw):
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    for a in APOSTROPHES:
        s = s.replace(a, "'")
    s = translit(s)
    s = re.sub(r"\s+", " ", s).strip(" ,.;-")
    if not s:
        return ""
    # so'zlarni to'g'ri bosh harf bilan
    words = []
    for w in s.split():
        wl = w.lower()
        if wl in ("o'g'li", "o'gli", "ogli", "ugli", "o'g'il"):
            words.append("o'g'li")
        elif wl in ("qizi", "kizi", "qzi"):
            words.append("qizi")
        else:
            words.append(w[0].upper() + w[1:] if len(w) > 1 else w.upper())
    return " ".join(words)


def is_junk_name(name):
    """Ism yaroqsizmi? (juda qisqa yoki harf yo'q)"""
    letters = re.sub(r"[^A-Za-z]", "", name)
    return len(letters) < 3


def core_candidates(digits):
    """9 xonali ehtimoliy 'core' raqamlar ro'yxati (eng ishonchli birinchi)."""
    n = len(digits)
    c = []
    if n == 9:
        c = [digits]
    elif n == 12 and digits.startswith("998"):
        c = [digits[3:]]
    elif n == 10:
        if digits[0] == "8":
            c.append(digits[1:])          # 8 = eski trunk prefiks
        c += [digits[:9], digits[-9:]]
    elif n == 11:
        c = [digits[-9:], digits[2:11]]
    elif n == 18:
        if digits.startswith("998"):
            c = [digits[3:12]]            # bitta uzun raqam (ko'pincha soxta)
        else:
            c = [digits[:9], digits[9:18]]  # ikkita raqam birikkan
    elif n > 12 and digits.startswith("998"):
        c = [digits[3:12]]
    elif n >= 9:
        c = [digits[-9:], digits[:9]]
    return c


def valid_core(core):
    return (len(core) == 9 and core[0] != "0" and core[:2] in VALID_PREFIX)


def is_fake(core):
    """Abonent qismi (oxirgi 7 raqam) hammasi bir xil bo'lsa -> soxta."""
    return len(set(core[2:])) == 1


def normalize_phone(raw):
    """(asosiy_raqam, ikkinchi_raqam, soxtami) qaytaradi. Yaroqsiz bo'lsa (None, None, ...)."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None, None, False
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return None, None, False

    valids = []
    for core in core_candidates(digits):
        if valid_core(core) and core not in valids:
            valids.append(core)

    if not valids:
        return None, None, False
    primary = valids[0]
    if is_fake(primary):
        return None, None, True            # soxta -> tekshirishga
    second = "+998" + valids[1] if len(valids) > 1 and not is_fake(valids[1]) else None
    return "+998" + primary, second, False


# ============ EXCEL BEZAK ============
def style_sheet(ws, n_rows, header_color, phone_col=2):
    head_fill = PatternFill("solid", fgColor=header_color)
    head_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols = ws.max_column
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1, min_col=1, max_col=ncols):
        for i, c in enumerate(row, start=1):
            c.border = border
            if i == phone_col:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = "@"
            else:
                c.alignment = Alignment(vertical="center")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    if ncols >= 3:
        ws.column_dimensions["C"].width = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + ncols)}1"


def main():
    if not Path(IN_FILE).exists():
        raise SystemExit(f"Fayl topilmadi: {IN_FILE}")

    df = pd.read_excel(IN_FILE, header=0).iloc[:, :2]
    df.columns = ["fish", "phone"]
    jami = len(df)

    toza, shubhali = [], []
    korilgan = set()
    stats = {"dublikat": 0, "soxta": 0, "raqamsiz": 0, "junk_ism": 0, "ikkinchi_raqam": 0}

    for _, r in df.iterrows():
        name = clean_name(r["fish"])
        primary, second, fake = normalize_phone(r["phone"])

        asl = "" if pd.isna(r["phone"]) else str(r["phone"]).strip()

        # yaroqsiz holatlar -> tekshirishga
        if primary is None:
            if fake:
                stats["soxta"] += 1
            elif not asl:
                stats["raqamsiz"] += 1
            shubhali.append([name, asl, "soxta raqam" if fake else "raqam yo'q/xato"])
            continue
        if is_junk_name(name):
            stats["junk_ism"] += 1
            shubhali.append([name, primary, "ism yaroqsiz"])
            continue

        # asosiy raqam
        if primary in korilgan:
            stats["dublikat"] += 1
        else:
            korilgan.add(primary)
            toza.append([name, primary])

        # ikkinchi raqam (bitta katakda 2 ta bo'lsa) ham saqlanadi
        if second and second not in korilgan:
            korilgan.add(second)
            toza.append([name, second])
            stats["ikkinchi_raqam"] += 1

    # ---------- Excel ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "Toza"
    ws.append(["F.I.Sh", "Telefon raqami"])
    for row in toza:
        ws.append(row)
    style_sheet(ws, len(toza), "16A34A")

    ws2 = wb.create_sheet("Tekshirish kerak")
    ws2.append(["F.I.Sh", "Asl raqam", "Sabab"])
    for row in shubhali:
        ws2.append(row)
    style_sheet(ws2, len(shubhali), "DC2626")

    ws3 = wb.create_sheet("Hisobot")
    rep = [
        ["Ko'rsatkich", "Qiymat"],
        ["O'qilgan qatorlar", jami],
        ["Toza (noyob) kontaktlar", len(toza)],
        ["  shundan 2-raqam (katakdagi)", stats["ikkinchi_raqam"]],
        ["Olib tashlangan dublikat", stats["dublikat"]],
        ["Tekshirish kerak (jami)", len(shubhali)],
        ["  - soxta raqam", stats["soxta"]],
        ["  - raqam yo'q/xato", stats["raqamsiz"]],
        ["  - ism yaroqsiz", stats["junk_ism"]],
    ]
    for row in rep:
        ws3.append(row)
    style_sheet(ws3, len(rep) - 1, "2563EB", phone_col=0)
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 14

    wb.save(OUT_FILE)

    # ---------- konsol hisoboti ----------
    print("=" * 48)
    print("            MUKAMMAL TOZALASH HISOBOTI")
    print("=" * 48)
    for k, v in rep[1:]:
        print(f"  {k:<32} {v}")
    print(f"  {'Natija fayli':<32} {OUT_FILE}")
    print("-" * 48)
    print("  Namuna (toza):")
    for row in toza[:6]:
        print(f"    {row[0]:<40} {row[1]}")


if __name__ == "__main__":
    main()
