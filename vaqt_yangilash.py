# -*- coding: utf-8 -*-
"""
game_dev_edu_list_toza.xlsx faylidagi "Created at" ustunini
2024-yil 1-dekabrdan 2025-yil 3-aprelgacha oraliqda qayta taqsimlaydi.

Sanalar ID (ro'yxatdan o'tish) tartibi bo'yicha o'sib boradi, vaqt 08:00-22:00 oralig'ida.
Boshqa ustunlar va bezak tegilmaydi (faqat Created at qiymatlari o'zgaradi).

Ishga tushirish:
    .venv\\Scripts\\python.exe vaqt_yangilash.py
"""
import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

FILE = "game_dev_edu_list_toza.xlsx"
START = datetime(2024, 12, 1, 8, 0, 0)
END = datetime(2025, 4, 3, 22, 0, 0)
DATE_FMT = "yyyy-mm-dd hh:mm:ss"

random.seed(42)


def make_timestamps(n):
    """n ta o'sib boradigan tasodifiy vaqt (START..END, 08:00-22:00)."""
    total = (END - START).total_seconds()
    times = []
    for f in sorted(random.random() for _ in range(n)):
        t = START + timedelta(seconds=f * total)
        t = t.replace(hour=random.randint(8, 21),
                      minute=random.randint(0, 59),
                      second=random.randint(0, 59), microsecond=0)
        times.append(t)
    times.sort()
    return times


def update_sheet(ws):
    # sarlavhadan ustun indekslarini topamiz
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    if "Created at" not in headers:
        return 0
    ca_col = headers["Created at"]
    id_col = headers.get("ID")

    rows = list(range(2, ws.max_row + 1))
    # ID bo'yicha tartiblaymiz (eng kichik ID -> eng erta sana)
    if id_col:
        def id_key(r):
            v = ws.cell(row=r, column=id_col).value
            try:
                return (0, float(v))
            except (TypeError, ValueError):
                return (1, 0)
        rows.sort(key=id_key)

    stamps = make_timestamps(len(rows))
    for r, ts in zip(rows, stamps):
        c = ws.cell(row=r, column=ca_col)
        c.value = ts
        c.number_format = DATE_FMT
    return len(rows)


def main():
    if not Path(FILE).exists():
        raise SystemExit(f"Fayl topilmadi: {FILE}. Avval: python dublikat.py")

    wb = load_workbook(FILE)
    jami = 0
    for name in wb.sheetnames:
        cnt = update_sheet(wb[name])
        if cnt:
            print(f"  '{name}' varag'i: {cnt} ta Created at yangilandi")
            jami += cnt
    wb.save(FILE)

    print("-" * 44)
    print(f"  Oraliq:  {START:%Y-%m-%d}  ->  {END:%Y-%m-%d}")
    print(f"  Jami yangilangan: {jami}")
    print(f"  Fayl saqlandi:    {FILE}")


if __name__ == "__main__":
    main()
