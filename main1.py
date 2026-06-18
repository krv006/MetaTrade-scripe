# -*- coding: utf-8 -*-
"""
650 ta odamning REAL ko'rinishdagi ma'lumotini (F.I.Sh + O'zbekiston telefon raqami)
Faker yordamida yaratib, Excel (.xlsx) faylga saqlaydi.

Ustunlar:
    1) Ism Familiya Sharifi   ->  masalan: Alisher Uzoqov Kamron o'g'li
    2) Telefon raqami         ->  masalan: +998901078383

Ishga tushirish:
    .venv\\Scripts\\python.exe main1.py
"""
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

fake = Faker()
Faker.seed(2026)

JAMI = 650
OUT_FILE = "mijozlar_650.xlsx"

# ============ REAL O'ZBEKCHA ISMLAR ============
ERKAK_ISMLAR = [
    "Alisher", "Akmal", "Aziz", "Azizbek", "Bekzod", "Botir", "Bobur", "Behruz", "Davron",
    "Dilshod", "Doston", "Diyor", "Eldor", "Elyor", "Farrux", "Firdavs", "Furqat", "Jahongir",
    "Javohir", "Jasur", "Jamshid", "Kamron", "Komil", "Laziz", "Murod", "Mirzo", "Nodir",
    "Nurbek", "Otabek", "Oybek", "Qodir", "Quvonch", "Ravshan", "Rustam", "Rasul", "Sardor",
    "Sanjar", "Sherzod", "Sherali", "Shoxrux", "Sirojiddin", "Said", "Temur", "Ulug'bek",
    "Umid", "Xusan", "Xurshid", "Yusuf", "Zafar", "Zohid", "Bahodir", "Ibrohim", "Islom",
    "Asror", "Anvar", "Asadbek", "Mardon", "Muzaffar", "Nizom", "Olim", "Ramz", "Sardorbek",
    "Shavkat", "Tohir", "Valijon", "Yodgor", "Zafarbek", "Abror", "Akbar", "Bahrom",
    "Davlat", "Eshmurod", "Fayzulla", "G'ayrat", "Habib", "Iskandar", "Jaloliddin", "Kamoliddin",
]

AYOL_ISMLAR = [
    "Dilnoza", "Gulnora", "Madina", "Malika", "Nargiza", "Nilufar", "Sevara", "Shahnoza",
    "Zarina", "Zebo", "Feruza", "Kamola", "Lola", "Mohira", "Nodira", "Oysha", "Rayhona",
    "Sabina", "Sevinch", "Shoira", "Umida", "Yulduz", "Zilola", "Dildora", "Gulchehra",
    "Hilola", "Iroda", "Kumush", "Maftuna", "Nozima", "Ozoda", "Ra'no", "Sabohat", "Charos",
    "Dilfuza", "Gavhar", "Husniya", "Komila", "Laylo", "Mavluda", "Nafisa", "Oygul", "Saodat",
    "Munisa", "Robiya", "Shahzoda", "Tursunoy", "Visola", "Xadicha", "Yorqinoy", "Zuhra",
    "Aziza", "Barno", "Dilrabo", "Farangiz", "Gulbahor", "Halima", "Iqbol", "Jamila", "Kibriyo",
]

# ============ REAL O'ZBEKCHA FAMILIYALAR (negiz) ============
FAMILIYA_NEGIZ = [
    "Karimov", "Rashidov", "Yusupov", "Uzoqov", "Rahimov", "Aliyev", "Ergashev", "Yo'ldoshev",
    "Ismoilov", "Sobirov", "To'xtayev", "Qodirov", "Nazarov", "Saidov", "Mirzayev", "Xolmatov",
    "Tursunov", "Bobojonov", "Jo'rayev", "Mahmudov", "Sodiqov", "Usmonov", "Hakimov", "Tojiboyev",
    "Abdullayev", "Sharipov", "G'aniyev", "Murodov", "Hamidov", "Salimov", "Yoqubov", "Niyozov",
    "Davlatov", "Sultonov", "To'rayev", "Berdiyev", "Eshonqulov", "Nematov", "Rustamov", "Sattorov",
    "Umarov", "Vohidov", "Xudoyberdiyev", "Yodgorov", "Zokirov", "Abdurahmonov", "Begmatov",
    "Choriyev", "Egamov", "Fayziyev", "Halilov", "Islomov", "Jalilov", "Komilov", "Latipov",
]

ARAB = "0123456789"
# Real O'zbekiston mobil operator prefikslari (Beeline, Ucell, Mobiuz, Uzmobile, Humans...)
PREFIKSLAR = ["90", "91", "93", "94", "95", "97", "98", "99", "88", "33", "77", "50", "20"]


def familiya(negiz, ayol):
    """Ayol uchun -ov/-ev -> -ova/-eva."""
    if ayol and (negiz.endswith("ov") or negiz.endswith("ev")):
        return negiz + "a"
    return negiz


def telefon():
    """Real prefiks bilan: +998XXYYYYYYY (9 raqam)"""
    prefiks = fake.random_element(PREFIKSLAR)
    raqam = "".join(fake.random_element(ARAB) for _ in range(7))
    return f"+998{prefiks}{raqam}"


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Mijozlar"

    # --- sarlavha (header) ---
    ws.append(["Ism Familiya Sharifi", "Telefon raqami"])

    head_fill = PatternFill("solid", fgColor="2563EB")
    head_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 3):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # --- ma'lumotlar ---
    used_phones = set()
    erkak_suffix, ayol_suffix = "o'g'li", "qizi"
    for _ in range(JAMI):
        ayol = fake.boolean(chance_of_getting_true=45)  # ~45% ayol
        ism = fake.random_element(AYOL_ISMLAR if ayol else ERKAK_ISMLAR)
        fam = familiya(fake.random_element(FAMILIYA_NEGIZ), ayol)
        ota = fake.random_element(ERKAK_ISMLAR)          # ota ismi har doim erkak ismi
        suffix = ayol_suffix if ayol else erkak_suffix
        fish = f"{ism} {fam} {ota} {suffix}"

        # noyob telefon
        tel = telefon()
        while tel in used_phones:
            tel = telefon()
        used_phones.add(tel)

        ws.append([fish, tel])

    # --- jadval ko'rinishi ---
    for row in ws.iter_rows(min_row=2, max_row=JAMI + 1, min_col=1, max_col=2):
        row[0].border = border
        row[0].alignment = Alignment(vertical="center")
        row[1].border = border
        row[1].alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A2"          # sarlavhani muzlatish
    ws.auto_filter.ref = "A1:B1"    # filtr

    wb.save(OUT_FILE)
    print(f"Tayyor! {JAMI} ta yozuv -> {OUT_FILE}")
    print("Namuna:")
    for r in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        print(f"  {r[0]:<40} {r[1]}")


if __name__ == "__main__":
    main()
