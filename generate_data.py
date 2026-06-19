# -*- coding: utf-8 -*-
"""
game_dev_edu_list formatida 100 ta REAL ko'rinishdagi yozuvni Faker bilan generatsiya qiladi.

Asosiy e'tibor: ISM va TELEFON RAQAM generatsiyasi (real o'zbekcha ismlar + O'zbekiston prefikslari).
Barcha ustunlar: Full Name, Gender, Birth Date, Email, Direction, Phone, Language,
                 Experience, Study or Staff, Online or Offline, Location, Created at.

Natija:  game_dev_generated.xlsx

Ishga tushirish:
    .venv\\Scripts\\python.exe generate_data.py
"""
import re
from datetime import datetime, timedelta

from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

fake = Faker()
Faker.seed(2025)

JAMI = 100
OUT_FILE = "game_dev_generated.xlsx"

# ============ ISM/FAMILIYA ============
ERKAK = ["Doston", "Javlonbek", "Abdulbosit", "Asadbek", "Davronbek", "Muhammad", "Shahriyor",
         "Otajon", "Muslim", "Khikmat", "Ibrokhim", "Ozodbek", "Muhammadmirzo", "Husan",
         "Anvarjon", "Faridun", "Behzod", "Bakhrom", "Diyorbek", "Xamid", "Abdumalik", "Davlat",
         "Alisher", "Bekzod", "Sardor", "Sherzod", "Jasur", "Temur", "Aziz", "Bobur", "Eldor",
         "Farrux", "Jahongir", "Kamron", "Nodir", "Otabek", "Rustam", "Sanjar", "Ulug'bek",
         "Xusan", "Zafar", "Akmal", "Bahodir", "Diyor", "Firdavs", "Islom", "Jamshid", "Komil"]
AYOL = ["Muslima", "Sevinch", "Mavluda", "Mohira", "Feruza", "Gulmira", "Sevara", "Nilufar",
        "Madina", "Malika", "Nargiza", "Dilnoza", "Zarina", "Sabina", "Shahnoza", "Umida",
        "Yulduz", "Kamola", "Nodira", "Charos", "Maftuna", "Oygul", "Saodat", "Munisa",
        "Robiya", "Dildora", "Iroda", "Komila", "Laylo", "Nafisa", "Zilola", "Aziza", "Barno"]
FAMILIYA = ["Qudratov", "Goziev", "Mirzajonov", "Botirov", "Normuminova", "Xudoyberganov",
            "Bozorboyev", "Sotvoldiyeva", "Abduvaliyev", "Allaberganov", "Abdurakhmonov",
            "Haitov", "Ochilov", "Pulatova", "Abdusattorov", "Tashmirzayeva", "Abduxamitov",
            "Erkinov", "Mamasoatova", "Komiljonov", "Xalillayeva", "Sadikov", "Karimov",
            "Rashidov", "Yusupov", "Rahimov", "Ergashev", "Ismoilov", "Qodirov", "Nazarov",
            "Saidov", "Mirzayev", "Tursunov", "Mahmudov", "Usmonov", "Hakimov", "Abdullayev",
            "Sharipov", "Murodov", "Hamidov", "Salimov", "Niyozov", "Sultonov", "Berdiyev"]

# ============ TANLOV RO'YXATLARI (sample asosida) ============
DIRECTIONS = (["Unity + C# dasturlash"] * 5 + ["3D Modelling"] * 3 +
              ["Game Designer"] * 2 + ["Game Studio biznesi"] * 2)
LANGUAGES = ["O'zbek tili", "O'zbek tili", "O'zbek tili", "Ingliz tili, O'zbek tili",
             "Rus tili, Ingliz tili", "O'zbek tili, Ingliz tili", "Rus tili",
             "Rus tili, O'zbek tili", "Ingliz tili", "Ingliz tili, Rus tili"]
EXPERIENCE = (["0 dan o'rganmoqchiman"] * 5 + ["Boshlang'ich ma'lumotga egaman"] * 3 +
              ["O'rta darajada ma'lumotga egaman"] * 2)
STUDY = ["Computer engineering", "Webster University", "Digital University",
         "New Uzbekistan University", "Amity University Tashkent", "TIUE", "AKITA TTPU",
         "Tramplin IT Academy", "Najot Ta'lim", "PDP Academy", "Renesans ta'lim universiteti",
         "Tashkent Turin Politexnika Universiteti", "Toshkent davlat pedagogika universiteti",
         "Hech qayerda ta'lim olmayman va ishlamayman", "Inha University", "TATU",
         "Westminster University", "Milliy universitet", "1-son IT maktab"]
LOCATIONS = (["Toshkent shahri"] * 6 + ["Toshkent viloyati"] * 2 + ["Samarqand viloyati"] * 2 +
             ["Andijon viloyati", "Namangan viloyati", "Farg'ona viloyati", "Xorazm viloyati",
              "Surxondaryo viloyati", "Sirdaryo viloyati", "Buxoro viloyati", "Jizzax viloyati",
              "Qashqadaryo viloyati", "Navoiy viloyati", "Qoraqalpog'iston Respublikasi"])
PREFIX = ["90", "91", "93", "94", "95", "97", "98", "99", "88", "33", "77", "50", "20"]
DOMAINS = ["gmail.com"] * 8 + ["mail.ru", "icloud.com"]

START = datetime(2024, 12, 1, 8, 0, 0)
END = datetime(2025, 4, 3, 22, 0, 0)


def latin(s):
    """Email uchun: apostrof/maxsus belgilarni olib tashlaydi, lotin qoldiradi."""
    s = s.replace("'", "").replace("'", "").lower()
    return re.sub(r"[^a-z]", "", s)


def make_phone(used):
    while True:
        p = f"+998{fake.random_element(PREFIX)}{''.join(fake.random_element('0123456789') for _ in range(7))}"
        if p not in used:
            used.add(p)
            return p


def make_email(first, last, used):
    f, l = latin(first), latin(last)
    while True:
        style = fake.random_int(0, 4)
        if style == 0:
            e = f"{f}{l}{fake.random_int(1, 999)}"
        elif style == 1:
            e = f"{f}.{l}"
        elif style == 2:
            e = f"{l}{f}{fake.random_int(10, 99)}"
        elif style == 3:
            e = f"{f}{fake.random_int(1, 9999)}"
        else:
            e = f"{f}_{l}"
        email = f"{e}@{fake.random_element(DOMAINS)}"
        if email not in used:
            used.add(email)
            return email


def fam_gender(last, ayol):
    """Familiyani jinsga moslaydi: erkak -> -ov/-ev, ayol -> -ova/-eva."""
    base = last[:-1] if (last.endswith("ova") or last.endswith("eva")) else last  # ayol -> erkak negiz
    if ayol and (base.endswith("ov") or base.endswith("ev")):
        return base + "a"
    return base


def make_name(first, last):
    """Sampledagi kabi turli tartib."""
    r = fake.random_int(0, 99)
    if r < 55:
        return f"{first} {last}"          # Ism Familiya
    elif r < 80:
        return f"{last} {first}"          # Familiya Ism
    elif r < 90:
        return first                      # faqat ism
    else:
        ota = fake.random_element(ERKAK)
        return f"{first} {ota} o'g'li {last}"  # patronimika bilan


def make_birthdate():
    y = fake.random_int(1996, 2010)
    m = fake.random_int(1, 12)
    d = fake.random_int(1, 28)
    return f"{y:04d}-{m:02d}-{d:02d}"


def make_created_times(n):
    total = (END - START).total_seconds()
    times = []
    for fr in sorted(fake.random.random() for _ in range(n)):
        t = (START + timedelta(seconds=fr * total)).replace(
            hour=fake.random_int(8, 21), minute=fake.random_int(0, 59),
            second=fake.random_int(0, 59))
        times.append(t)
    times.sort()
    return times


def main():
    used_phone, used_email = set(), set()
    times = make_created_times(JAMI)
    rows = []
    for i in range(JAMI):
        ayol = fake.boolean(chance_of_getting_true=40)
        first = fake.random_element(AYOL if ayol else ERKAK)
        last = fam_gender(fake.random_element(FAMILIYA), ayol)
        rows.append([
            i + 1,
            make_name(first, last),
            "Ayol" if ayol else "Erkak",
            make_birthdate(),
            make_email(first, last, used_email),
            fake.random_element(DIRECTIONS),
            make_phone(used_phone),
            fake.random_element(LANGUAGES),
            fake.random_element(EXPERIENCE),
            fake.random_element(STUDY),
            fake.random_element(["online", "offline"]),
            fake.random_element(LOCATIONS),
            times[i].strftime("%Y-%m-%d %H:%M:%S"),
        ])

    # ---------- Excel ----------
    headers = ["№", "Full Name", "Gender", "Birth Date", "Email", "Direction", "Phone",
               "Language", "Experience", "Study or Staff", "Online or Offline", "Location", "Created at"]
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(headers)
    for r in rows:
        ws.append(r)

    # bezak
    head_fill = PatternFill("solid", fgColor="2563EB")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    phone_col = headers.index("Phone") + 1
    for row in ws.iter_rows(min_row=2, max_row=JAMI + 1, min_col=1, max_col=len(headers)):
        for i, c in enumerate(row, start=1):
            c.border = border
            c.alignment = Alignment(vertical="center")
            if i == phone_col:
                c.number_format = "@"
    widths = {"№": 5, "Full Name": 26, "Gender": 8, "Birth Date": 13, "Email": 32,
              "Direction": 22, "Phone": 17, "Language": 22, "Experience": 28,
              "Study or Staff": 30, "Online or Offline": 16, "Location": 22, "Created at": 20}
    for idx, name in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(name, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}1"
    wb.save(OUT_FILE)

    print(f"Tayyor! {JAMI} ta yozuv -> {OUT_FILE}")
    print("Namuna:")
    for r in rows[:6]:
        print(f"  {r[0]:>3}  {r[1]:<28} {r[6]:<15} {r[4]}")


if __name__ == "__main__":
    main()
