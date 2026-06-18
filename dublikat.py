# -*- coding: utf-8 -*-
"""
"game_dev_edu_list (2).xlsx" dan dublikatlarni olib tashlaydi.

Mezon: bir xil ODAM = EMAIL yoki TELEFON mos kelsa (telefon formati e'tiborga olinmaydi).
Saqlanadi: har bir odamning BIRINCHI (eng eski, Created at bo'yicha) yozuvi.

Ma'lumotning o'zi o'zgartirilmaydi — faqat ortiqcha qatorlar olib tashlanadi.

Natija:  game_dev_edu_list_toza.xlsx
    * "data"            -> tozalangan (noyob) ro'yxat
    * "Olib tashlangan" -> o'chirilgan dublikatlar (yo'qotmaslik uchun)

Ishga tushirish:
    .venv\\Scripts\\python.exe dublikat.py
"""
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

IN_FILE = "game_dev_edu_list (2).xlsx"
OUT_FILE = "game_dev_edu_list_toza.xlsx"
SHEET = "data"


def norm_email(x):
    s = str(x).strip().lower()
    return "e:" + s if "@" in s else None


def norm_phone(x):
    d = re.sub(r"\D", "", str(x))
    if len(d) >= 12 and d.startswith("998"):
        d = d[3:12]
    elif len(d) >= 9:
        d = d[-9:]
    return "p:" + d if len(d) == 9 else None


class UF:
    def __init__(self, n):
        self.p = list(range(n))

    def find(self, a):
        while self.p[a] != a:
            self.p[a] = self.p[self.p[a]]
            a = self.p[a]
        return a

    def union(self, a, b):
        self.p[self.find(a)] = self.find(b)


def style(ws, n_rows, n_cols, header_color, phone_idx=None):
    head_fill = PatternFill("solid", fgColor=header_color)
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1, min_col=1, max_col=n_cols):
        for i, c in enumerate(row):
            c.border = border
            c.alignment = Alignment(vertical="center")
            if phone_idx is not None and i == phone_idx:
                c.number_format = "@"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=n_cols).column_letter}1"


def main():
    if not Path(IN_FILE).exists():
        raise SystemExit(f"Fayl topilmadi: {IN_FILE}")

    df = pd.read_excel(IN_FILE, sheet_name=SHEET)
    n = len(df)

    # eng eski birinchi qolishi uchun Created at bo'yicha tartiblaymiz
    if "Created at" in df.columns:
        df["_ca"] = pd.to_datetime(df["Created at"], errors="coerce")
        df = df.sort_values("_ca", kind="stable").reset_index(drop=True)

    # union-find: email yoki telefon mos kelsa birlashtiramiz
    uf = UF(n)
    buckets = {}
    for i in range(n):
        for k in (norm_email(df.at[i, "Email"]), norm_phone(df.at[i, "Phone"])):
            if k:
                buckets.setdefault(k, []).append(i)
    for ids in buckets.values():
        for j in ids[1:]:
            uf.union(ids[0], j)

    df["_grp"] = [uf.find(i) for i in range(n)]

    # har guruhdan birinchisi (eng eski) -> toza; qolgani -> olib tashlangan
    keep_mask = ~df.duplicated(subset="_grp", keep="first")
    toza = df[keep_mask].drop(columns=["_ca", "_grp"], errors="ignore")
    olib = df[~keep_mask].drop(columns=["_ca", "_grp"], errors="ignore")

    # ID bo'yicha qayta tartiblash (asl ko'rinish uchun)
    if "ID" in toza.columns:
        toza = toza.sort_values("ID", kind="stable")
        olib = olib.sort_values("ID", kind="stable")

    # ---------- yozish ----------
    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as xw:
        toza.to_excel(xw, sheet_name="data", index=False)
        olib.to_excel(xw, sheet_name="Olib tashlangan", index=False)

    # bezak
    wb = load_workbook(OUT_FILE)
    cols = list(toza.columns)
    phone_idx = cols.index("Phone") if "Phone" in cols else None
    style(wb["data"], len(toza), len(cols), "16A34A", phone_idx)
    style(wb["Olib tashlangan"], len(olib), len(cols), "DC2626", phone_idx)
    # ustun kengligi
    widths = {"Full Name": 24, "Email": 30, "Direction": 24, "Phone": 18,
              "Language": 22, "Experience": 26, "Study or Staff": 28,
              "Location": 18, "Created at": 24}
    for ws in (wb["data"], wb["Olib tashlangan"]):
        for idx, name in enumerate(cols, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(name, 14)
    wb.save(OUT_FILE)

    # ---------- hisobot ----------
    print("=" * 46)
    print("        DUBLIKAT OLIB TASHLASH HISOBOTI")
    print("=" * 46)
    print(f"  Mezon:                 Email yoki Telefon")
    print(f"  Saqlangan nusxa:       birinchisi (eng eski)")
    print(f"  Asl qatorlar:          {n}")
    print(f"  Toza (noyob) odamlar:  {len(toza)}")
    print(f"  Olib tashlangan:       {len(olib)}")
    print(f"  Natija fayli:          {OUT_FILE}")


if __name__ == "__main__":
    main()
