# -*- coding: utf-8 -*-
"""
HAQIQIY ma'lumotni anonimlashtiradi: faqat "Full Name" va "Phone" ustunlarini
Faker bilan generatsiya qilingan qiymatlarga almashtiradi.

Boshqa BARCHA ustun (Gender, Birth Date, Email, Direction, Language, Experience,
Study or Staff, Online or Offline, Location, Created at) REAL holatda qoladi.

Ism mavjud "Gender" ustuniga MOS keladi (Erkak -> erkak ismi, Ayol -> ayol ismi).
Telefonlar +998XXXXXXXXX formatida va noyob.

Manba:   game_dev_edu_list_toza.xlsx   (tozalangan, dublikatsiz)
Natija:  game_dev_anonim.xlsx
    * asl fayl o'zgartirilmaydi.

Ishga tushirish:
    .venv\\Scripts\\python.exe anonim.py
"""
from pathlib import Path

from faker import Faker
from openpyxl import load_workbook

from generate_data import ERKAK, AYOL, FAMILIYA, fam_gender, PREFIX

fake = Faker()
Faker.seed(777)

IN_FILE = "game_dev_edu_list_toza.xlsx"
OUT_FILE = "game_dev_anonim.xlsx"
LIMIT = None   # None = barcha qatorlar; masalan 100 desangiz faqat birinchi 100 ta


def make_phone(used):
    while True:
        p = "+998" + fake.random_element(PREFIX) + "".join(fake.random_element("0123456789") for _ in range(7))
        if p not in used:
            used.add(p)
            return p


def make_name(gender):
    ayol = str(gender).strip().lower() == "ayol"
    first = fake.random_element(AYOL if ayol else ERKAK)
    last = fam_gender(fake.random_element(FAMILIYA), ayol)
    r = fake.random_int(0, 99)
    if r < 55:
        return f"{first} {last}"
    elif r < 80:
        return f"{last} {first}"
    elif r < 90:
        return first
    return f"{first} {fake.random_element(ERKAK)} o'g'li {last}"


def process_sheet(ws, used):
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    name_col = headers.get("Full Name")
    phone_col = headers.get("Phone")
    gender_col = headers.get("Gender")
    if not name_col or not phone_col:
        return 0

    last_row = ws.max_row
    if LIMIT:
        last_row = min(ws.max_row, LIMIT + 1)

    cnt = 0
    for r in range(2, last_row + 1):
        gender = ws.cell(row=r, column=gender_col).value if gender_col else "Erkak"
        ws.cell(row=r, column=name_col).value = make_name(gender)
        pc = ws.cell(row=r, column=phone_col)
        pc.value = make_phone(used)
        pc.number_format = "@"
        cnt += 1
    return cnt


def main():
    if not Path(IN_FILE).exists():
        raise SystemExit(f"Fayl topilmadi: {IN_FILE}. Avval: python dublikat.py")

    wb = load_workbook(IN_FILE)
    used = set()
    jami = 0
    for name in wb.sheetnames:
        c = process_sheet(wb[name], used)
        if c:
            print(f"  '{name}' varag'i: {c} ta Ism+Telefon almashtirildi")
            jami += c
    wb.save(OUT_FILE)

    print("-" * 46)
    print(f"  Faqat o'zgardi:  Full Name + Phone")
    print(f"  Real qoldi:      qolgan barcha ustun")
    print(f"  Jami yozuv:      {jami}")
    print(f"  Natija fayli:    {OUT_FILE}")


if __name__ == "__main__":
    main()
